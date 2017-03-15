import requests
import sys
import json
import openpyxl
import re
from talker import Talker
from IPython.display import HTML, display
from ipywidgets import *
from tabulate import tabulate

class Parser(object):
	"""This class makes many get requests and parses the data"""
	def __init__(self, talker, site):
		super(Parser, self).__init__()
		self.talker = talker
		self.site = site

	def getElement(self, idd):
		return self.talker.getElementById(self.site, idd)

	def getTable(self, idd, editable=False):
		# fix this once we find the things that are in common here
		if editable:
			self.getEditableTable(idd)
		else:
			self.getUneditableTable(idd)


	def getUneditableTable(self, idd):
		#ETO Analysis - "_18_0_5_3a40149_1487976881089_349552_19724"
		#PAF Structual Capability - "_18_0_5_3a40149_1487274522343_479963_15422"
		#D3 - "_18_0_5_3a40149_1486497151977_210118_16085"
		#Sample Data Table - "_18_0_5_3a40149_1486149200614_128977_15535"
		tableAddress = "_18_0_5_3a40149_1487976881089_349552_19724"
		table = self.talker.getElementById("mbppguidex", tableAddress)
		table = json.dumps(table)
		data = json.loads(table)

		instanceSpecifications = data["elements"][0]["specialization"]["instanceSpecificationSpecification"]["string"]
		titles, body = instanceSpecifications.split("body", 1)
		headers = titles.split("<p>")
		headerList = []
		for header in headers:
		    headerList.append(header.split("<\/p>")[0])
		headerList.pop(0)
		#print(headerList)

		rows = body.split("name")
		rows.pop(0)
		rowList = []
		for row in rows:
			columnList = []
			columns = row.split('"source":"')
			columns.pop(0)
			for column in columns:
				itemID = column.split('"')[0]
				variable = self.talker.getElementById("mbppguidex", itemID)
				variable = json.dumps(variable)
				variable  = json.loads(variable)
				if "value" in variable["elements"][0]["specialization"]:
					columnList.append(variable["elements"][0]["specialization"]["value"][0]["double"])
				else:
					columnList.append(variable["elements"][0]["name"])
			rowList.append(columnList)
		print(tabulate(rowList, headers=headerList))

	def getEditableTable(self, idd):
		caption = Label('editable table here', disabled=True)

		button = widgets.Button(description="Update MMS", layout=Layout(position='bottom'))
		button.on_click(self.on_button_clicked)

		display(caption, button)

	def on_button_clicked(self, b):
		# save the updated values in the table to a json object
		# sends the json back to server
		print("MMS Updated")
	def updateTableExcel(self, idd):
		table = self.talker.getElementById("mbppguidex", "_18_0_5_3a40149_1487274522343_479963_15422")
		table = json.dumps(table)
		data = json.loads(table)

		fileName = "paf_capability_1.xlsx"
		sheetName = input('\nSheet Name: ')
		tableStart = input('Cell ID of start of Table: ')
		excelTable = excelParser(fileName, sheetName, tableStart)

		instanceSpecifications = data["elements"][0]["specialization"]["instanceSpecificationSpecification"]["string"]
		titles, body = instanceSpecifications.split("body", 1)
		headers = titles.split("<p>")
		headerList = []
		for header in headers:
		    headerList.append(header.split("<\/p>")[0])
		headerList.pop(0)

		rows = body.split("name")
		rows.pop(0)
		rowList = []
		rowPlace = 1

		for row in rows:
		    columnList = []
		    columnPlace = 0
		    columns = row.split('"source":"')
		    columns.pop(0)
		    for column in columns:
		        itemID = column.split('"')[0]
		        variable = self.talker.getElementById("mbppguidex", itemID)
		        variable = json.dumps(variable)
		        variable  = json.loads(variable)
		        if "value" in variable["elements"][0]["specialization"]:
		            if variable["elements"][0]["specialization"]["value"][0]["double"] != excelTable[rowPlace][columnPlace]:
						#Number of floating points for equivalence can be further clarified
		                variable["elements"][0]["specialization"]["value"][0]["double"] = "%.5f" % excelTable[rowPlace][columnPlace]
		                self.talker.postElementById("mbppguidex", variable)
		        else:
		            if variable["elements"][0]["name"] != excelTable[rowPlace][columnPlace]:
		                variable["elements"][0]["name"] = excelTable[rowPlace][columnPlace]
		                self.talker.postElementById("mbppguidex", variable)
		        columnPlace = columnPlace + 1
		    rowPlace = rowPlace + 1

	def excelParser(fileName, sheetName, tableStart):
	    wb = openpyxl.load_workbook(fileName, data_only=True)
	    sheet = wb.get_sheet_by_name(sheetName)
	    tableStart = tableStart
	    column = ord(re.split('(\d+)', tableStart)[0]) - 64
	    row = int(re.split('(\d+)', tableStart)[1])
	    rowCheck = row
	    columnCheck = column
	    columns = 0
	    rows = 0
	    rowList = []
	    while sheet.cell(row = rowCheck, column = column).value != None or rowCheck == row:
	        rowCheck = rowCheck + 1
	        rows = rows + 1
	    while sheet.cell(row = row, column = columnCheck).value != None or columnCheck == column:
	        columns = columns + 1
	        columnCheck = columnCheck + 1
	    for xrow in range(0, rows):
	        columnList = []
	        for ycolumn in range(0, columns):
	            columnList.append(sheet.cell(row = row + xrow, column = column + ycolumn).value)
	        rowList.append(columnList)
	    return rowList


class Value(object):
    def __init__(self, value):
        self.value = value
    #this may not be fair to assume that lengths will always be 3 or 5, but worked in this case
    # def isNumeric(self):
    #     return len(self.value["elements"][0]["specialization"]["type"]) in ["LiteralReal", "LiteralInteger"]
    def isNumeric(self):
        return self.value["elements"][0]["specialization"]["type"] in ["LiteralReal", "LiteralInteger"]
    def isName(self):
        return len(self.value["elements"][0]["specialization"]) < 5
    #gets name or number depending on type based on above assumptions
    def getName(self):
        return self.value["elements"][0]["name"]
    def getNumber(self):
        return self.value["elements"][0]["specialization"]["value"][0]["double"]

class Number(object):
    iD = ""
    num = 0

    def __init__(self, arg):
        self.iD = arg["elements"][0]["sysmlid"]
        self.num = arg["elements"][0]["specialization"]["value"][0]["double"]

class Title(object):
    iD = ""
    title = ""

    def __init__(self, arg):
        self.iD = arg["elements"][0]["sysmlid"]
        self.title = arg["elements"][0]["name"]

class Line(object):
    title = ""
    column1 = ""
    column2 = ""

    def __init__(self, title, column1, column2):
        self.title = title
        self.column1 = column1
        self.column2 = column2
